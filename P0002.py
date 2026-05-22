
prog = 'P0002'
nome = 'Itensfoj_Criacao'
versao = 'v1'

#--------------------------------------------------------------------------------------------------
# imports
from os import getcwd, chdir, listdir, path, scandir, remove, system
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from urllib import request
import subprocess
from shutil import copyfile

#--------------------------------------------------------------------------------------------------
# título
system('cls') or None

print(f'\n{nome}\n({versao} - {prog})')


# -------------------------------------------------------------------------------------------------
# momento atual
print('\n -- 02.01 - Momento atual -------------------------------------------------------------------------')

momento_atual = datetime.now()
ano_atual = momento_atual.isocalendar()[0]

dia = momento_atual.day
mes = momento_atual.month
ano = momento_atual.year

print(f'{dia}/{mes}/{ano}')


# -------------------------------------------------------------------------------------------------
# informação de semana baseada na atual
print('\n -- 02.02 informação de semana baseada na atual ---------------------------------------------------')

print(f'ano_atual -> {ano_atual}')
print(f'- Ano atual: {ano_atual}')

semana_atual = momento_atual.isocalendar()[1]
print(f'- Semana atual: {semana_atual}')

mes_atual = momento_atual.month
print(f'- Mês atual: {mes_atual}')

# -------------------------------------------------------------------------------------------------
# quantas semanas estão prontas para cada ano
print('\n -- 02.03 - busca quantas semanas estão prontas para cada ano -------------------------------------')

caminho_ano_itensfoj = f'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_10_Custos_Referencia\\Itensfoj'
print(f'caminho_ano_itensfoj -> {caminho_ano_itensfoj}')

print(getcwd())


for ano in range(2018, ano_atual + 1):
    ano1 = listdir(f'{caminho_ano_itensfoj}\\{ano}')
    #print(f'ano1 -> {ano1}')
    ano1 = len(ano1)
    print(f'- Semanas em {ano}: {ano1}')
    if ano < ano_atual:
        item = 52
    else:
        item = semana_atual
    faltam_ano = item - int(ano1)
    print(f'- para {ano} faltam {faltam_ano} semanas')


# -------------------------------------------------------------------------------------------------
# recebe informação de qual o ano que será analisado
print('\n -- 02.04 - recebe informação de qual o ano que será analisado ------------------------------------')

ano = input('\nQual o ano? ')


#--------------------------------------------------------------------------------------------------
# traz data atual no modelo iso calendar
print('\n -- 02.05 - traz data atual no modelo iso calendar ------------------------------------------------')

data_atual = datetime.now()
data_atual = date.isocalendar(data_atual)
print(f'data_atual (aaaa, w, dia_da_semana) -> {data_atual}')


# -------------------------------------------------------------------------------------------------
# diz para no ano atual ir apenas até a semana atual
print('\n-- 02.06 - diz para no ano atual ir apenas até a semana atual -------------------------------------')

if int(ano) == int(ano_atual):
    n_semana = int(semana_atual)
else:
    n_semana = 52

print(f'para o ano de {ano}, verifica até a semana - > {n_semana}')


#--------------------------------------------------------------------------------------------------
# busca de quais semanas já tem e quais faltam
print('\n-- 02.07 - busca de quais semanas já tem e quais faltam -------------------------------------------')

chdir(f'{caminho_ano_itensfoj}\\{ano}')
print(getcwd())
arquivos = listdir()
pasta1 = list(scandir())
qtd_arquivos = len(listdir())
#print(f'arquivos -> {arquivos}')

print(f'qtd de itensfoj na pasta do ano -> {qtd_arquivos}')

#print(arquivos)

semana_faltando = []

for semana1 in range(1, n_semana + 1):
    if semana1 < 10:
        semana2 = f'0{semana1}'
    else:
        semana2 = semana1
    #print(f'semana2 -> {semana2}')
    semana3 = str(f'Itensfoj_{ano}W{semana2}.xlsx')
    #print(f'semana3 -> {semana3}')
    if semana3 in arquivos:
        try:
            item6 = pasta1[int(semana2) - 1].stat().st_size
            #print(f'item6 -> {item6}')
        except:
            item6 = 1
        if item6 < 100000:
            print(f'{ano}W{semana2}_tamanho->{item6}_sem informação')
        else:
            pass
        #print(f'{ano}W{semana2} - OK')
    else:
        faltando = [ano, semana2]
        semana_faltando.append(faltando)
        print(f'{ano}W{semana2} - criar arquivo')


qtd_semana_faltando = len(semana_faltando)
print(f'qtd_semana_faltando -> {qtd_semana_faltando}')
print(f'semana_faltando -> {semana_faltando}')


#--------------------------------------------------------------------------------------------------
# definir se abrir algum itensfoj ou se gerar algum novo
print('\n-- 02.08 - definir se abrir algum itensfoj ou se gerar algum novo ---------------------------------')


semana = input("\nQual a semana? ('s' para sair)\n")
while semana != 's':
    semana = int(semana)
    if semana < 10:
        semana = int(f'0{semana}')
    else:
        pass
    #print(semana)
    print('\ng - gerar apenas um itensfoj')
    print('v - varios (completar lacunas do ano)')
    print('a - abrir algum itensfoj')
    print('s - sair')
    comando = input('\no que vc deseja fazer? ')
    if comando == 'a':
        chdir(f'C:\\Users\\jk65078\\Box Sync\\Laptop\\29.10_Custos_Referencia\\Itensfoj\\{ano}')
        print(getcwd())
        system(f"Itensfoj_{ano}W0{semana}.xlsx")
    elif comando == 'g':
        break
    elif comando == 'v':
        break
    elif comando == 's':
        exit()
    else:
        print('repete')
else:
    exit()


# -------------------------------------------------------------------------------------------------
# busca informações da planilha de fornecedores
print('\n -- 02.09 - busca informações da planilha de fornecedores -----------------------------------------')

caminho_fornecedores = 'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_18_Fornecedores'
arquivo_fornecedores = 'Listagem_fornecedores.xlsx'

wb2 = load_workbook(f'{caminho_fornecedores}\\{arquivo_fornecedores}')
ws2 = wb2['Sheet1']

qtd_linhas_fornecedores = ws2.max_row

lista_fornecedores = []
for i in range(6, qtd_linhas_fornecedores + 1):
    Num_Fornecedor = int(ws2[f'B{i}'].value)
    if Num_Fornecedor != None:
        Fornecedor = ws2[f'E{i}'].value
        Commodity = ws2[f'H{i}'].value
        lista_fornecedores.append([Num_Fornecedor, Fornecedor, Commodity])

print(f'lista_fornecedores -> {lista_fornecedores}')
qtd_lista_fornecedores = len(lista_fornecedores)
print(f'qtd_lista_fornecedores -> {qtd_lista_fornecedores}')

wb2.close()




# -------------------------------------------------------------------------------------------------
# logica de repetição do itensfoj
print('\n -- 02.10 - lógica de repetição do itensfoj  ------------------------------------------------------')

if comando == 'v':
    pass
else:
    qtd_semana_faltando = 1

for i in range(0, qtd_semana_faltando):
    if comando == 'v':
        ano = int(semana_faltando[i][0])
        semana = int(semana_faltando[i][1])
    else:
        pass

    #-----data-----------------------------------------------------------------------------------------
    # definição de data - inicio
    print('\n -- 02.11 - definição de data - inicio ------------------------------------------------------------')
    
    ano = int(ano)
    print(f'ano -> {ano}')
    print(type(ano))

    if ano == 2017:
        data = datetime(year = 2016, month = 12, day = 31)
    elif ano == 2018:
        data = datetime(year = 2017, month = 12, day = 30)
    elif ano == 2019:
        data = datetime(year = 2018, month = 12, day = 29)
    elif ano == 2020:
        data = datetime(year = 2019, month = 12, day = 28)
    else:
        print('erro')
    print(f'data -> {data}')
    data_wd = data.weekday()
    print(f'data_wd -> {data_wd}')

    lista_datas_semana = []
    indices = [1, 2, 3, 4, 5, 6, 7]
    for i in range(0, 7):
        data1 = (data + timedelta(days = indices[i] + (7 * (semana - 1))))
        lista_datas_semana.append(data1)
        data2 = f'{data1.day}/{data1.month}/{data1.year}'

    
    qtd_lista_datas_semana = len(lista_datas_semana)
    print(f'qtd_lista_datas_semana -> {qtd_lista_datas_semana}')
    print(f'lista_datas_semana -> {lista_datas_semana}')



    # -------------------------------------------------------------------------------------------------
    # itensfoj semanal introdução - inicio
    print('\n -- 02.12 - itensfoj semanal introdução - inicio --------------------------------------------------')

    caminho_ano_download = f'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_10_Custos_Referencia\\Itensfoj\\download\\{ano}'
    pasta_ano = chdir(caminho_ano_download)

    print('\n...rodando...\n')

    lista1 = listdir(pasta_ano)
    qtd_lista1 = len(lista1)
    print(f'qtd_lista1 -> {qtd_lista1}')
    #print(lista1)

    lista2 = []
    for item in range(0, qtd_lista_datas_semana):
        dia = lista_datas_semana[item].day
        if dia < 10:
            dia = f'0{dia}'
        else:
            pass

        mes = lista_datas_semana[item].month
        if mes < 10:
            mes = f'0{mes}'
        else:
            pass

        data10 = f'{lista_datas_semana[item].year}{mes}{dia}' 
        print(f'item1 -> {data10}')
        for item2 in range(0, qtd_lista1):
            item3 = lista1[item2].split('_')
            print(f'item3 -> {item3}')
            try:
                item4 = str(item3[3])
            except:
                item4 = str(item3[2])
            #print(f'item4 -> {item4}')
            item5 = item4.split('.')[0]
            #print(f'item5 -> {item5}')
            if data10 == item5:
                #print(f'item2')
                lista2.append(lista1[item2])
            else:
                pass

    print(f'lista2 -> {lista2}')
    qtd_lista2 = len(lista2)

    if qtd_lista2 == 0:
        print('não há informações para esta semana...')
    else:
        pass

    print(f'qtd_lista2 -> {qtd_lista2}')


    # -------------------------------------------------------------------------------------------------
    # buscando informacoes do arquivo
    print('\n -- 02.13 - buscando informações do arquivo -------------------------------------------------------')


    caminho = f'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_10_Custos_Referencia\\Itensfoj\\download\\{ano}'
    chdir(caminho)
    print(getcwd())



    indice = 0
    lista = []
    while indice < len(lista2):
        arquivo_origem = lista2[indice]
        print(f'arquivo_origem -> {arquivo_origem}')
        lista1 = []
        with open(arquivo_origem, encoding = 'latin-1') as arquivo:
            arquivo = list(arquivo)
            qtd_linhas = len(arquivo)
            print(f'qtd_linhas -> {qtd_linhas}')
            for info in range(0, qtd_linhas):
                info1 = arquivo[info]
                info1 = info1.split(';')
                qtd_info1 = len(info1)
                if qtd_info1 < 43:
                    print(f'qtd_info1 (itens na linha) -> {len(info1)}')
                else:
                    if info1[0] == 'Fornecedor':
                        pass
                    else:
                        lista1.append(info1)
                        pass
                    pass

            
        lista = lista + lista1
        celula1 = str(lista[0][2]).strip()
        indice = indice + 1
    try:
        celula1 = str(lista[0][0]).strip()
        print(f'celula1 (primeiro item do primeiro arquivo) -> {celula1}')
    except:
        pass

    print('\n...rodando...\n')

    print('informações carregadas dos arquivos txt')


    # ---------------------------------------------------------------------------------------------
    # inicio do processamento das informacoes - inicio
    print('\n -- 02.14 - inicio do processamento das informações - inicio --------------------------------------')

    # primeira linha
    primeira_linha = 0
    #print(f'\nid da primeira linha: {primeira_linha}')
    #print(f'{lista[0][2][2].strip()} - {lista[0][4][2].strip()}')

    # última linha da lista
    ultima_linha = len(lista) - 1
    #print(f'\nr da ultima linha: {ultima_linha}')
    quantidade_total_de_linhas = len(lista)
    #print(f'quantidade total de linhas: {quantidade_total_de_linhas}')
    #print(f'{lista[ultima_linha][2][2].strip()} - {lista[ultima_linha][4][2].strip()}')


    #print(f'{lista[ultima_linha - primeira_linha][2][2].strip()}')
    import os
    #print(os.getcwd())

    os.chdir('C:\\Users\\jk65078\\Box Sync\\Laptop\\29.26 - Modelos')


    # inicion do processamento das informações - fim
    #--------------------------------------------------------------------------------------------------
    #abrindo arquivo e identificando a semana - inicio
    print('\n -- 02.15 - abrindo arquivo e identificando a semana - inicio -------------------------------------')

    import openpyxl
    from openpyxl import load_workbook

    caminho_modelos = 'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_26_Modelos'
    arquivo_modelo_itensfoj = '48_Itensfoj_0000W00_modelo.xlsx'

    wb = load_workbook(f'{caminho_modelos}\\{arquivo_modelo_itensfoj}')
    ws1 = wb['Sheet1']

    #semana = senama + 1
    # correção necessária conforme observação no primeiro uso de 'semana', onde é corrigida a diferença entre o número de seman AGCO e o Python

    if semana < 10:
        semana = f'0{semana}'
    else:
        pass

    ws1['B1'] = 'itensfoj'
    ws1['C1'] = f'Itensfoj_{ano}W{semana}'
    ws1['B2'] = 'data referência'
    print(f'lista_datas_semana -> {lista_datas_semana}')
    dia2 = lista_datas_semana[1].day
    if dia2 < 10:
        dia2 = f'0{dia2}'
    else:
        pass
    mes2 = lista_datas_semana[1].month
    if mes2 < 10:
        mes2 = f'0{mes2}'
    else:
        pass
    ano2 = lista_datas_semana[1].year

    ws1['C2'] = f'{dia2}/{mes2}/{ano2}'
    ws1['B3'] = 'data criação'
    dia3 = momento_atual.day
    if dia3 < 10:
        dia3 = f'0{dia3}'
    else:
        pass
    mes3 = momento_atual.month
    if mes3 < 10:
        mes3 = f'0{mes3}'
    else:
        pass
    ano3 = momento_atual.year

    ws1['C3'] = f'{dia3}/{mes3}/{ano3}'

    
    # -------------------------------------------------------------------------------------------------
    # filtrando e colando PN - inicio
    print('\n -- 02.16 - filtrando e colando PN - inicio -------------------------------------------------------')
    

    ws1['B5'] = 'PN'
    for num in range(0, quantidade_total_de_linhas):
        celula1 = str(lista[num][2]).strip()
        if celula1[-2:] == '-1':
            celula = celula1[:-2]
        else:
            if celula1[-2:] == '-0':
                celula = celula1[:-2]
            else:
                celula = celula1
        ws1[f'B{num + 6}'] = celula
        num = num + 1

    try:
        PN_ref = celula
        pass
    except:
        pass


    # -------------------------------------------------------------------------------------------------
    # descrição do item - inicio
    print('\n -- 02.17 - descrição do item - inicio ------------------------------------------------------------')


    ws1['C5'] = 'Descrição do Item'
    for num in range(0, quantidade_total_de_linhas):
        celula = lista[num][3].strip()
        ws1[f'C{num + 6}'] = celula
        num = num + 1

    try:
        #print(f'Descrição -> {celula}')
        pass
    except:
        pass

    ws1['D5'] = 'Chave1'
    for num in range(0, quantidade_total_de_linhas):
        celula1 = str(lista[num][2]).strip()
        if celula1[-2:] == '-1':
            celula = celula1[:-2]
        else:
            if celula1[-2:] == '-0':
                celula = celula1[:-2]
            else:
                celula = celula1
        celula = celula \
            .__add__('_')\
            .__add__(str(lista[num][22]).strip()) \
            .__add__('_')\
            .__add__(str(int(lista[num][0])))
        ws1[f'D{num + 6}'] = celula
        #print(celula)
        num = num + 1

    try:
        #print(f'chave1 -> {celula}')
        pass
    except:
        pass

    ws1['E5'] = 'Chave2'
    for num in range(0, quantidade_total_de_linhas):
        celula1 = str(lista[num][2]).strip()
        if celula1[-2:] == '-1':
            celula = celula1[:-2]
        else:
            if celula1[-2:] == '-0':
                celula = celula1[:-2]
            else:
                celula = celula1
        celula = celula \
            .__add__('_')\
            .__add__(str(lista[num][22]).strip())
        ws1[f'E{num + 6}'] = celula
        #print(celula)
        num = num + 1

    try:
        #print(f'chave2 -> {celula}')
        pass
    except:
        pass

    ws1['F5'] = 'Chave3'
    for num in range(0, quantidade_total_de_linhas):
        celula1 = str(lista[num][2]).strip()
        if celula1[-2:] == '-1':
            celula = celula1[:-2]
        else:
            if celula1[-2:] == '-0':
                celula = celula1[:-2]
            else:
                celula = celula1
        celula = celula \
            .__add__('_')\
            .__add__(str(int(lista[num][0])))
        ws1[f'F{num + 6}'] = celula
        num = num + 1


    # teste para melhorar o loop de volume anual
    ws1['G5'] = 'Volume Anual'
    volume_total = {}
    for num in range(0, quantidade_total_de_linhas):
        celula1 = str(lista[num][2]).strip()
        if celula1[-2:] == '-1':
            PN = celula1[:-2]
        else:
            if celula1[-2:] == '-0':
                PN = celula1[:-2]
            else:
                PN = celula1
        celula2 = round(float(lista[num][8].replace(',','.')), 0)
        valor = volume_total.get(f'{PN}')
        if valor == None:
            valor1 = 0
        else:
            valor1 = volume_total.get(f'{PN}')
        valor = valor1 + celula2
        volume_total.update({f'{PN}': valor})

        ws1[f'G{num + 6}'] = volume_total.get(f'{PN}')
        volume_splitado = round(float(lista[num][9].replace(',','.')), 0)
        if volume_total.get(f'{PN}') == 0:
            celula = 0
            #print(celula)
        else:
            celula = volume_splitado / volume_total.get(f'{PN}')
            #print(celula)
        ws1[f'H{num + 6}'] = celula
        #print(celula)
        num = num + 1
        if num == int(quantidade_total_de_linhas / 10):
            print(f'10% - ok {quantidade_total_de_linhas / 10}')
        if num == int((quantidade_total_de_linhas / 10) * 2):
            print(f'20% - ok {((quantidade_total_de_linhas / 10) * 2)}')
        if num == int((quantidade_total_de_linhas / 10) * 3):
            print(f'30% - ok {((quantidade_total_de_linhas / 10) * 3)}')
        if num == int((quantidade_total_de_linhas / 10) * 4):
            print(f'40% - ok {((quantidade_total_de_linhas / 10) * 4)}')
        if num == int((quantidade_total_de_linhas / 10) * 5):
            print(f'50% - ok {((quantidade_total_de_linhas / 10) * 5)}')
        if num == int((quantidade_total_de_linhas / 10) * 6):
            print(f'60% - ok {((quantidade_total_de_linhas / 10) * 6)}')
        if num == int((quantidade_total_de_linhas / 10) * 7):
            print(f'70% - ok {((quantidade_total_de_linhas / 10) * 7)}')
        if num == int((quantidade_total_de_linhas / 10) * 8):
            print(f'80% - ok {((quantidade_total_de_linhas / 10) * 8)}')
        if num == int((quantidade_total_de_linhas / 10) * 9):
            print(f'90% - ok {((quantidade_total_de_linhas / 10) * 9)}')
        if num == int((quantidade_total_de_linhas / 10) * 10):
            print(f'100% - ok {quantidade_total_de_linhas}')



    try:
        volume = volume_total.get(f'{PN_ref}')
    except:
        pass

    ws1['H5'] = 'Split'

    # split é calculado junto aos volumes total e splitado no item logo acima

    ws1['I5'] = 'Custo Unitário'
    for num in range(0, quantidade_total_de_linhas):
        Moeda = str(lista[num][29]) # M
        valor_na_moeda = lista[num][27] # O
        valor_internado = lista[num][30] # Q
        valor_contrato = lista[num][7]# L
        if Moeda == 'BRL':
            celula = valor_na_moeda
        else:
            if float(valor_internado.replace(',','.')) == 0:
                celula = 0
            else:
                celula = round(((float(valor_na_moeda.replace(',','.')) / float(valor_internado.replace(',','.'))) * float(valor_contrato.replace(',','.'))), 4)
        ws1[f'I{num + 6}'] = celula
        #print(celula)
        num = num + 1

    try:
        #print(f'custo unit -> {celula}')
        pass
    except:
        pass

    ws1['J5'] = 'Num_Fornecedor'
    ws1['K5'] = 'Fornecedor'
    for num in range(0, quantidade_total_de_linhas):
        celula = int(lista[num][0])
        ws1[f'J{num + 6}'] = celula
        ws1[f'K{num + 6}'] = lista[num][11]
        for j in range(0, qtd_lista_fornecedores):
            if celula == int(lista_fornecedores[j][0]):
                ws1[f'K{num + 6}'] = lista_fornecedores[j][1]
            else:
                pass
        num = num + 1


    ws1['L5'] = 'Site'
    for num in range(0, quantidade_total_de_linhas):
        celula = str(lista[num][22]).strip()
        ws1[f'L{num + 6}'] = celula
        #print(celula)
        num = num + 1

    try:
        #print(f'Site -> {celula}')
        pass
    except:
        pass

    ws1['M5'] = 'Preço Moeda Orig'
    for num in range(0, quantidade_total_de_linhas):
        Moeda = str(lista[num][29])
        if Moeda == 'BRL':
            celula = float(lista[num][7].replace(',','.'))
        else:
            celula = float(lista[num][30].replace(',','.'))
        ws1[f'M{num + 6}'] = celula
        num = num + 1


    ws1['N5'] = 'Moeda'
    for num in range(0, quantidade_total_de_linhas):
        celula = str(lista[num][29])
        ws1[f'N{num + 6}'] = celula
        num = num + 1

    ws1['O5'] = 'Rate'
    for num in range(0, quantidade_total_de_linhas):
        Moeda = str(lista[num][29])
        Moeda_companhia = float(lista[num][27].replace(',','.'))  # O
        Valor_internado = float(lista[num][30].replace(',','.'))  # Q
        if Moeda == 'BRL':
            celula = 1
        else:
            if Valor_internado == 0:
                celula = 0
            else:
                celula = round((Moeda_companhia / Valor_internado), 2)
        ws1[f'O{num + 6}'] = celula
        num = num + 1

    ws1['P5'] = 'Preço BRL/PES'
    for num in range(0, quantidade_total_de_linhas):
        Moeda = str(lista[num][29])
        if Moeda == 'BRL':
            celula1 = float(lista[num][7].replace(',','.'))
        else:
            celula1 = float(lista[num][30].replace(',','.'))
        Moeda_companhia = float(lista[num][27].replace(',','.'))  # O
        Valor_internado = float(lista[num][30].replace(',','.'))  # Q
        if Moeda == 'BRL':
            celula2 = 1
        else:
            if Valor_internado == 0:
                celula2 = 1
            else:
                celula2 = round((Moeda_companhia / Valor_internado), 4)
        celula = round((celula1 * celula2), 2)
        ws1[f'P{num + 6}'] = celula
        num = num + 1

    ws1['Q5'] = 'Impostos'
    for num in range(0, quantidade_total_de_linhas):
        Moeda = str(lista[num][29])
        Moeda_companhia = float(lista[num][27].replace(',','.'))  # O
        valor_contrato = float(lista[num][7].replace(',','.'))  # L
        if Moeda == 'BRL':
            if valor_contrato == 0:
                celual = 0
            else:
                celula = 1 - (Moeda_companhia / valor_contrato)
        else:
            celula = 0
        ws1[f'Q{num + 6}'] = celula
        num = num + 1

    try:
        #print(f'impostos -> {celula}')
        pass
    except:
        pass

    lista = []


    wb.save(f'C:\\Users\\jk65078\\Box Sync\\Laptop\\29_10_Custos_Referencia\\Itensfoj\\{ano}\\Itensfoj_{ano}W{semana}.xlsx')
    wb.save(f'V:\\SCA_Supplier_Cost_Analysis\\01_SCA\\10_Team\\Jorge\\Itensfoj\\{ano}\\Itensfoj_{ano}W{semana}.xlsx')
    wb.close()



# -------------------------------------------------------------------------------------------------
# busca volumes de 'budget' para inclusão no itensfoj gerado
print('\n -- 02.18 - busca volumes de "budget" para inclusao no itensfoj gerado ----------------------------')


import P0041


# -------------------------------------------------------------------------------------------------
# criar itensfoj_semana_atual
print('\n -- 02.19 - criar itenfoj_semana_atual ------------------------------------------------------------')

comando = input('atualiza itensfoj_semana_atual? (s - sim)')

if comando == 's':
    lista_ano1 = []
    lista_ano = listdir(caminho_ano_itensfoj)
    print(f'lista_ano -> {lista_ano}')
    for i in range(0, len(lista_ano)):
        print(f'lista_ano[i] -> {lista_ano[i]}')
        valor = lista_ano[i]
        try:
            valor = int(valor)
            lista_ano1.append(valor)
        except:
            pass
    print(f'lista_ano1 -> {lista_ano1}')
    ano1 = lista_ano1[-1]
    print(f'ano1 -> {ano1}')

    lista_itensfoj = listdir(f'{caminho_ano_itensfoj}\\{ano1}')
    print(f'lista_itensfoj -> {lista_itensfoj}')

    ultimo_itensfoj = lista_itensfoj[-1]
    print(f'ultimo_itensfoj -> {ultimo_itensfoj}')

    itensfoj_semana_atual_origem = f'{caminho_ano_itensfoj}\\{ano1}\\{ultimo_itensfoj}'
    itensfoj_semana_atual_destino = f'{caminho_ano_itensfoj}\\Itensfoj_sem_atual.xlsx'
    copyfile(itensfoj_semana_atual_origem, itensfoj_semana_atual_destino)
else:
    pass

